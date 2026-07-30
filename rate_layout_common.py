"""Shared helpers for Global Runners rate layout builders."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from project_paths import INPUT_DIR, OUTPUT_DIR, PROCESSING_DIR, ensure_workspace_dirs
from supplier_name_lookup import map_fred_supplier_names

OCEAN_SOURCE_SHEET = "Ocean"
AIR_SOURCE_SHEET = "Air"
PROCESSING_GLOB = "combined_*.xlsx"
EXCEL_SUFFIXES = {".xlsx", ".xls", ".xlsm"}

COST_NAME_ROW = 1
APPLY_IF_ROW = 2
RATE_BY_ROW = 3
COLUMN_HEADER_ROW = 4
DATA_START_ROW = 5

HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
BLUE_FILL = PatternFill("solid", fgColor="BDD7EE")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
BOLD = Font(bold=True)
NORMAL = Font()
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

CURRENCY_COLUMN = "Quoting Currency"

FTL_PRE_CARRIAGE_COL = "FTL Pre-carriage (Flat fee)"
LTL_MIN_PRE_CARRIAGE_COL = "LTL Pre-carriage (Minimum fee)"
LTL_PUNIT_PRE_CARRIAGE_COL = "LTL Pre-carriage (Cost per Chargeable KG)"
LTL_PUNIT_PRE_CARRIAGE_AIR_COL = "LTL Pre-carriage ( per KG)"
DG_SURCHARGE_COL = "Pre-Carriage DG Surcharge [FIXED %]"

ROAD_PRECARRIAGE_SHIPMENT_COLUMNS = [
    "Type",
    "Second load",
    "Origin Country",
    "Port of loading POL",
    "Origin Base city",
    "Destination country",
    "Destination location id",
    "Destination Base City(only EI SG, DHL UK, DHL UX)",
    "Service",
    "Carrier Name",
    "Supplier Code",
    "Valid From",
    "Valid to",
]

ROAD_PRECARRIAGE_AIR_SHIPMENT_COLUMNS = [
    "Rate card",
    "Service Type",
    "Lane UID",
    "Origin Country",
    "Origin Base City",
    "Destination Country",
    "Service",
    "Supplier Name (Q)",
    "Carrier Name",
    "Destination Airport Code",
    "Destination City (for DSV US only)",
    "Valid from",
    "Valid to",
]

ROAD_PRECARRIAGE_SEA_BOLD_COLUMNS = frozenset(
    {
        "Second load",
        "Origin Country",
        "Port of loading POL",
        "Origin Base city",
        "Destination country",
        "Destination location id",
        "Destination Base City(only EI SG, DHL UK, DHL UX)",
        "Service",
        "Carrier Name",
        "Valid From",
        "Valid to",
    }
)

ROAD_PRECARRIAGE_AIR_BOLD_COLUMNS = frozenset(
    {
        "Origin Country",
        "Origin Base City",
        "Destination Country",
        "Service",
        "Carrier Name",
        "Destination Airport Code",
    }
)

OCEAN_SECOND_LOAD_MAP = {
    "LCL": "LTL/LCL",
    "FCL": "FCL/Multi_load",
}

AIR_TO_OCEAN_SERVICE_TYPE = {
    "MAWB": "FCL",
    "HAWB": "LCL",
}


@dataclass(frozen=True)
class CostValueColumn:
    header: str
    source_column: str | None = None
    percent_over_costs: bool = False


@dataclass(frozen=True)
class CostBlock:
    title: str
    apply_if: str
    rate_by: str
    value_columns: tuple[CostValueColumn, ...]
    match_equipment_type: str | None = None
    match_service_type: str | None = None
    match_carrier_groups: tuple[str, ...] | None = None
    match_arx_carrier_lane: bool = False
    match_dzs_aei_carrier_lane: bool = False
    match_supplier_name: str | None = None
    exclude_arx_carrier_lane: bool = False
    exclude_dzs_aei_carrier_lane: bool = False
    fill_when_min_equals_max: bool = False
    fill_when_min_not_equals_max: bool = False
    include_currency: bool = True


def cell_text(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def display_text(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(value).lstrip()


def capitalize_type_label(value: object) -> str:
    text = display_text(value)
    if not text:
        return ""
    return text[0].upper() + text[1:]


def trim_lane_uid_suffix(lane_uid: object) -> str:
    text = display_text(lane_uid)
    if not text:
        return ""
    upper = text.upper()
    for suffix in ("MAWB", "HAWB"):
        if upper.endswith(suffix):
            return text[: -len(suffix)]
    return text


DZS_AEI_SUPPLIER_NAME = "DZS AEI"


def format_date(value: object) -> str:
    if pd.isna(value):
        return ""
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return cell_text(value)
    return parsed.strftime("%d.%m.%Y")


def rate_value(value: object) -> float | None:
    if pd.isna(value):
        return None
    text = cell_text(value)
    if not text or text.lower() in {"on request", "n/a", "#n/a"}:
        return None
    try:
        return float(str(value).replace(",", "."))
    except (TypeError, ValueError):
        return None


def round_numeric_output(value: object) -> float | None:
    number = rate_value(value)
    if number is None:
        return None
    return round(float(number), 2)


def format_cell_value(value: object) -> object:
    rounded = round_numeric_output(value)
    if rounded is None:
        return None
    return rounded


def format_shipment_cell_value(value: object) -> object:
    if value == "" or pd.isna(value):
        return None
    rounded = round_numeric_output(value)
    if rounded is not None:
        return rounded
    text = display_text(value)
    return text or None


def apply_two_decimal_number_format(cell) -> None:
    if isinstance(cell.value, (int, float)) and not (
        isinstance(cell.value, float) and pd.isna(cell.value)
    ):
        cell.number_format = "0.00"


def format_percent_over_costs_value(value: object) -> object:
    """Source may be 20 (percent) or 0.2 (fraction); output always uses percent (20)."""
    number = rate_value(value)
    if number is None:
        return None
    if 0 < number < 1:
        number = number * 100
    return format_cell_value(number)


def apply_grouped_second_column_fill(
    shipment_df: pd.DataFrame,
    *,
    first_column: str,
    second_column: str,
    group_columns: tuple[str, ...],
) -> pd.DataFrame:
    if first_column not in shipment_df.columns or second_column not in shipment_df.columns:
        return shipment_df

    result = shipment_df.copy()
    normalized = result[list(group_columns)].map(
        lambda value: "" if pd.isna(value) else str(value).strip()
    )
    group_keys = normalized.agg("|".join, axis=1)
    for group_key in group_keys.unique():
        if not group_key:
            continue
        mask = group_keys == group_key
        if mask.sum() <= 1:
            continue
        result.loc[mask, second_column] = result.loc[mask, first_column]

    return result


def output_equipment_type(equipment_type: str) -> str:
    text = cell_text(equipment_type)
    if text.upper().endswith("DC"):
        return f"{text[:-2]}FT"
    return text


def sorted_equipment_types(ocean_df: pd.DataFrame) -> list[str]:
    if "Equipment Type" not in ocean_df.columns:
        return []

    equipment_types = {
        output_equipment_type(value)
        for value in ocean_df["Equipment Type"].dropna().tolist()
        if cell_text(value)
    }
    return sorted(equipment_types, key=lambda value: (value.upper() == "LCL", value))


def destination_location_id_from_code(code_value: object) -> str:
    text = cell_text(code_value)
    if not text:
        return ""
    before_dash = text.split("-", 1)[0].strip()
    if len(before_dash) <= 2:
        return ""
    return before_dash[2:]


def map_ocean_second_load(service_type: object) -> str:
    text = cell_text(service_type)
    return OCEAN_SECOND_LOAD_MAP.get(text.upper(), text)


def prepare_air_for_precarriage_costs(air_df: pd.DataFrame) -> pd.DataFrame:
    prepared = air_df.copy()
    prepared["Service Type"] = prepared["Service Type"].map(
        lambda value: AIR_TO_OCEAN_SERVICE_TYPE.get(cell_text(value).upper(), cell_text(value))
    )
    if (
        LTL_PUNIT_PRE_CARRIAGE_AIR_COL in prepared.columns
        and LTL_PUNIT_PRE_CARRIAGE_COL not in prepared.columns
    ):
        prepared[LTL_PUNIT_PRE_CARRIAGE_COL] = prepared[LTL_PUNIT_PRE_CARRIAGE_AIR_COL]
    return prepared


def map_air_service_type_to_ocean(service_type: object) -> str:
    text = cell_text(service_type).upper()
    return AIR_TO_OCEAN_SERVICE_TYPE.get(text, cell_text(service_type))


def build_road_precarriage_shipment_from_ocean(ocean_df: pd.DataFrame) -> pd.DataFrame:
    empty = pd.Series([""] * len(ocean_df), index=ocean_df.index, dtype="object")
    service = pd.Series(["Pre-carriage"] * len(ocean_df), index=ocean_df.index, dtype="object")

    return pd.DataFrame(
        {
            "Type": ocean_df["Service Type"],
            "Second load": ocean_df["Service Type"].map(map_ocean_second_load),
            "Origin Country": ocean_df["Origin Country Code"],
            "Port of loading POL": empty,
            "Origin Base city": ocean_df["Origin Base City"],
            "Destination country": ocean_df["Origin Country Code"],
            "Destination location id": ocean_df["Port of discharge POD (UNLOCODE)"].map(
                destination_location_id_from_code
            ),
            "Destination Base City(only EI SG, DHL UK, DHL UX)": ocean_df["Destination Base City"],
            "Service": service,
            "Carrier Name": map_fred_supplier_names(
                ocean_df["Origin Country Code"],
                ocean_df["Supplier"],
                transport_mode="ocean",
                destination_countries=ocean_df["Destination Country Code"],
            ),
            "Supplier Code": ocean_df["Supplier"],
            "Valid From": ocean_df["Valid from"].map(format_date),
            "Valid to": ocean_df["Valid to"].map(format_date),
        },
        columns=ROAD_PRECARRIAGE_SHIPMENT_COLUMNS,
    )


def build_road_precarriage_shipment_from_air(
    air_df: pd.DataFrame,
    *,
    rc_name: str,
) -> pd.DataFrame:
    empty = pd.Series([""] * len(air_df), index=air_df.index, dtype="object")
    service = pd.Series(["Pre-carriage"] * len(air_df), index=air_df.index, dtype="object")

    return pd.DataFrame(
        {
            "Rate card": rc_name,
            "Service Type": air_df["Service Type"],
            "Lane UID": air_df["Lane UID"].map(trim_lane_uid_suffix),
            "Origin Country": air_df["Origin Country Code"],
            "Origin Base City": air_df["Origin Base City"],
            "Destination Country": air_df["Origin Country Code"],
            "Service": service,
            "Supplier Name (Q)": air_df["Supplier Name (Q)"].map(display_text),
            "Carrier Name": map_fred_supplier_names(
                air_df["Origin Country Code"],
                air_df["Supplier Name (Q)"],
                transport_mode="air",
                destination_countries=air_df["Destination Country Code"],
            ),
            "Destination Airport Code": air_df["Destination Airport Code"].map(cell_text),
            "Destination City (for DSV US only)": empty,
            "Valid from": air_df["Valid from"].map(format_date),
            "Valid to": air_df["Valid to"].map(format_date),
        },
        columns=ROAD_PRECARRIAGE_AIR_SHIPMENT_COLUMNS,
    )


def build_road_precarriage_cost_blocks(
    source_df: pd.DataFrame,
    *,
    for_air: bool = False,
) -> list[CostBlock]:
    blocks: list[CostBlock] = []
    service_types = {
        cell_text(value).upper()
        for value in source_df["Service Type"].dropna().tolist()
        if cell_text(value)
    }

    if for_air:
        if "FCL" in service_types:
            blocks.append(
                CostBlock(
                    title="Local pre-carriage road freight (FTL: FCL)",
                    apply_if="Apply if: Service type equals 'FCL'",
                    rate_by="Rate by: Per shipment",
                    value_columns=(CostValueColumn("Flat", FTL_PRE_CARRIAGE_COL),),
                    match_service_type="FCL",
                )
            )
    else:
        for equipment_type in sorted_equipment_types(source_df):
            if equipment_type.upper() == "LCL":
                continue

            blocks.append(
                CostBlock(
                    title=f"Local pre-carriage road freight (FTL : FCL {equipment_type})",
                    apply_if=f"Apply if: Equipment type equals '{equipment_type}'",
                    rate_by="Rate by: Per shipment",
                    value_columns=(CostValueColumn("Flat", FTL_PRE_CARRIAGE_COL),),
                    match_equipment_type=equipment_type,
                )
            )

    if "FCL" in service_types:
        blocks.extend(
            [
                CostBlock(
                    title="Local pre-carriage road freight (LTL: FCL)",
                    apply_if="Apply if: Service type equals 'FCL'",
                    rate_by="Rate by: Weight/kg",
                    value_columns=(
                        CostValueColumn("MIN", LTL_MIN_PRE_CARRIAGE_COL),
                        CostValueColumn("p/unit", LTL_PUNIT_PRE_CARRIAGE_COL),
                        CostValueColumn("MAX", FTL_PRE_CARRIAGE_COL),
                    ),
                    match_service_type="FCL",
                    fill_when_min_not_equals_max=True,
                ),
                CostBlock(
                    title="Local pre-carriage road freight (LTL: FCL)",
                    apply_if="Apply if: Service type equals 'FCL'",
                    rate_by="Rate by: Per shipment",
                    value_columns=(CostValueColumn("Flat", FTL_PRE_CARRIAGE_COL),),
                    match_service_type="FCL",
                    fill_when_min_equals_max=True,
                ),
            ]
        )

    if "LCL" in service_types:
        blocks.extend(
            [
                CostBlock(
                    title="Local pre-carriage road freight (FTL: LCL)",
                    apply_if="Apply if: Service type equals 'LCL'",
                    rate_by="Rate by: Per shipment",
                    value_columns=(CostValueColumn("Flat", FTL_PRE_CARRIAGE_COL),),
                    match_service_type="LCL",
                ),
                CostBlock(
                    title="Local pre-carriage road freight (LTL: LCL)",
                    apply_if="Apply if: Service type equals 'LCL'",
                    rate_by="Rate by: Weight/kg",
                    value_columns=(
                        CostValueColumn("MIN", LTL_MIN_PRE_CARRIAGE_COL),
                        CostValueColumn("p/unit", LTL_PUNIT_PRE_CARRIAGE_COL),
                        CostValueColumn("MAX", FTL_PRE_CARRIAGE_COL),
                    ),
                    match_service_type="LCL",
                    fill_when_min_not_equals_max=True,
                ),
                CostBlock(
                    title="Local pre-carriage road freight (LTL: LCL)",
                    apply_if="Apply if: Service type equals 'LCL'",
                    rate_by="Rate by: Per shipment",
                    value_columns=(CostValueColumn("Flat", FTL_PRE_CARRIAGE_COL),),
                    match_service_type="LCL",
                    fill_when_min_equals_max=True,
                ),
            ]
        )

    blocks.append(
        CostBlock(
            title="Local pre-carriage DGR surcharge",
            apply_if="",
            rate_by="Rate by: % over costs",
            value_columns=(
                CostValueColumn(
                    "% over costs",
                    DG_SURCHARGE_COL,
                    percent_over_costs=True,
                ),
            ),
            include_currency=False,
        )
    )

    return blocks


def list_input_files() -> list[Path]:
    return [
        path
        for path in sorted(INPUT_DIR.iterdir())
        if path.is_file()
        and path.suffix.lower() in EXCEL_SUFFIXES
        and not path.name.startswith("~$")
    ]


def get_rate_card_name(rate_card_name: str | None = None) -> str:
    if rate_card_name:
        return rate_card_name
    files = list_input_files()
    if not files:
        raise FileNotFoundError(f"No Excel files found in: {INPUT_DIR}")
    return files[0].stem


def rate_card_name_from_processing_path(processing_path: Path) -> str | None:
    """Read source rate card name from combined_<name>_<timestamp>.xlsx."""
    import re

    stem = processing_path.stem
    if not stem.startswith("combined_"):
        return None
    rest = stem[len("combined_") :]
    match = re.match(r"^(.+)_\d{8}_\d{6}$", rest)
    if not match:
        return None
    return match.group(1)


def output_workbook_path(rate_card_name: str | None = None) -> Path:
    import re

    rc_name = re.sub(r'[\\/*?:\[\]]+', "_", get_rate_card_name(rate_card_name)).strip()
    return OUTPUT_DIR / f"{rc_name}_output.xlsx"


def latest_processing_file() -> Path:
    files = sorted(PROCESSING_DIR.glob(PROCESSING_GLOB))
    if not files:
        raise FileNotFoundError(
            f"No processing workbook found in {PROCESSING_DIR}. Run process_tabs.py first."
        )
    return files[-1]


def load_ocean_dataframe(processing_path: Path | None = None) -> pd.DataFrame:
    path = processing_path or latest_processing_file()
    workbook = pd.ExcelFile(path)
    sheet_name = next(
        (name for name in workbook.sheet_names if name.lower() == OCEAN_SOURCE_SHEET.lower()),
        None,
    )
    if sheet_name is None:
        raise ValueError(f"Sheet '{OCEAN_SOURCE_SHEET}' not found in {path.name}")
    return pd.read_excel(path, sheet_name=sheet_name)


def load_air_dataframe(processing_path: Path | None = None) -> pd.DataFrame:
    path = processing_path or latest_processing_file()
    workbook = pd.ExcelFile(path)
    sheet_name = next(
        (name for name in workbook.sheet_names if name.lower() == AIR_SOURCE_SHEET.lower()),
        None,
    )
    if sheet_name is None:
        raise ValueError(f"Sheet '{AIR_SOURCE_SHEET}' not found in {path.name}")
    return pd.read_excel(path, sheet_name=sheet_name)


def block_width(block: CostBlock) -> int:
    if block.include_currency:
        return 1 + len(block.value_columns)
    return len(block.value_columns)


def block_has_any_source_value(source_row: pd.Series, block: CostBlock) -> bool:
    for value_column in block.value_columns:
        if value_column.source_column is None:
            continue
        if rate_value(source_row.get(value_column.source_column)) is not None:
            return True
    return False


def row_matches_block(
    source_row: pd.Series,
    block: CostBlock,
    *,
    transport_mode: str | None = None,
    shipment_row: pd.Series | None = None,
    supplier_name_column: str = "Supplier name",
) -> bool:
    if block.match_supplier_name is not None:
        if shipment_row is None:
            return False
        return cell_text(shipment_row.get(supplier_name_column)) == block.match_supplier_name

    if block.match_equipment_type is not None:
        source_equipment = cell_text(source_row["Equipment Type"])
        if output_equipment_type(source_equipment) != block.match_equipment_type:
            return False
    if block.match_service_type is not None:
        if cell_text(source_row["Service Type"]).upper() != block.match_service_type.upper():
            return False

    if (
        block.match_carrier_groups is not None
        or block.match_arx_carrier_lane
        or block.exclude_arx_carrier_lane
        or block.exclude_dzs_aei_carrier_lane
    ):
        from carrier_groups import (
            carrier_group_for_origin,
            is_arx_carrier_supplier_name,
        )
        from supplier_name_lookup import lookup_fred_supplier_name

        supplier_column = "Supplier Name (Q)" if transport_mode == "air" else "Supplier"
        supplier_name = lookup_fred_supplier_name(
            source_row.get("Origin Country Code"),
            source_row.get(supplier_column),
            transport_mode=transport_mode,
        )
        arx_lane = is_arx_carrier_supplier_name(supplier_name)

        if block.match_arx_carrier_lane:
            return arx_lane
        if block.exclude_arx_carrier_lane and arx_lane:
            return False
        if block.exclude_dzs_aei_carrier_lane and shipment_row is not None:
            if cell_text(shipment_row.get(supplier_name_column)) == DZS_AEI_SUPPLIER_NAME:
                return False
        if block.match_carrier_groups is not None:
            carrier_group = carrier_group_for_origin(source_row.get("Origin Country Code"))
            if carrier_group not in block.match_carrier_groups:
                return False

        if (
            block.match_equipment_type is None
            and block.match_service_type is None
            and block.match_carrier_groups is None
        ):
            return False

    return True


def should_fill_block(
    source_row: pd.Series,
    block: CostBlock,
    *,
    transport_mode: str | None = None,
    shipment_row: pd.Series | None = None,
    supplier_name_column: str = "Supplier name",
) -> bool:
    if not row_matches_block(
        source_row,
        block,
        transport_mode=transport_mode,
        shipment_row=shipment_row,
        supplier_name_column=supplier_name_column,
    ):
        return False
    if not (block.fill_when_min_equals_max or block.fill_when_min_not_equals_max):
        return True

    min_value, max_value, _ = pre_carriage_min_max_values(source_row)
    if block.fill_when_min_equals_max:
        return min_equals_max(min_value, max_value)
    if block.fill_when_min_not_equals_max:
        return not min_equals_max(min_value, max_value)
    return True


def pre_carriage_min_max_values(ocean_row: pd.Series) -> tuple[float | None, float | None, float | None]:
    min_value = rate_value(ocean_row.get(LTL_MIN_PRE_CARRIAGE_COL))
    max_value = rate_value(ocean_row.get(FTL_PRE_CARRIAGE_COL))
    punit_value = rate_value(ocean_row.get(LTL_PUNIT_PRE_CARRIAGE_COL))
    return min_value, max_value, punit_value


def min_equals_max(min_value: float | None, max_value: float | None) -> bool:
    if min_value is None or max_value is None:
        return False
    return float(min_value) == float(max_value)


def style_header_cell(cell, *, bold: bool = False, center: bool = False) -> None:
    cell.font = BOLD if bold else NORMAL
    cell.fill = HEADER_FILL
    cell.alignment = CENTER if center else LEFT


def write_merged_header_row(
    ws: Worksheet,
    row_index: int,
    start_col: int,
    width: int,
    value: str,
) -> None:
    end_col = start_col + width - 1
    if width > 1:
        ws.merge_cells(
            start_row=row_index,
            start_column=start_col,
            end_row=row_index,
            end_column=end_col,
        )
    cell = ws.cell(row=row_index, column=start_col, value=value)
    style_header_cell(cell)


def write_cost_block_headers(ws: Worksheet, block: CostBlock, start_col: int) -> None:
    width = block_width(block)

    write_merged_header_row(ws, COST_NAME_ROW, start_col, width, block.title)
    write_merged_header_row(ws, APPLY_IF_ROW, start_col, width, block.apply_if)
    write_merged_header_row(ws, RATE_BY_ROW, start_col, width, block.rate_by)

    value_start = 1 if block.include_currency else 0
    if block.include_currency:
        currency_cell = ws.cell(row=COLUMN_HEADER_ROW, column=start_col, value="Currency")
        style_header_cell(currency_cell, center=True)

    for offset, value_column in enumerate(block.value_columns, start=value_start):
        cell = ws.cell(row=COLUMN_HEADER_ROW, column=start_col + offset, value=value_column.header)
        style_header_cell(cell, center=True)


def write_rate_value_cell(ws: Worksheet, row: int, column: int, value: object) -> None:
    if value is None:
        return
    cell = ws.cell(row=row, column=column, value=value)
    cell.number_format = "0.00"


def _normalize_row_cell_for_signature(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, (int, float)):
        number = float(value)
        if pd.isna(number):
            return ""
        return f"{round(number, 2):.2f}"
    return str(value).strip()


def highlight_fully_duplicate_lane_rows(
    ws: Worksheet,
    *,
    shipment_columns: list[str],
    last_data_column: int,
    duplicate_lane_columns: tuple[str, ...],
    fill: PatternFill | None = None,
) -> int:
    """Fill rows where the selected shipment columns match another row."""
    if last_data_column < 1 or ws.max_row < DATA_START_ROW or not duplicate_lane_columns:
        return 0

    from collections import defaultdict

    include_columns = [
        col_idx
        for col_idx, header in enumerate(shipment_columns, start=1)
        if header in duplicate_lane_columns
    ]

    rows_by_signature: dict[str, list[int]] = defaultdict(list)
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        signature = "\x1f".join(
            _normalize_row_cell_for_signature(ws.cell(row=row_idx, column=col_idx).value)
            for col_idx in include_columns
        )
        if not signature.replace("\x1f", ""):
            continue
        rows_by_signature[signature].append(row_idx)

    row_fill = fill or GREEN_FILL
    highlighted_rows = 0
    for row_indices in rows_by_signature.values():
        if len(row_indices) < 2:
            continue
        for row_idx in row_indices:
            for col_idx in range(1, last_data_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = row_fill
            highlighted_rows += 1
    return highlighted_rows


def write_matrix_sheet(
    ws: Worksheet,
    ocean_df: pd.DataFrame,
    shipment_df: pd.DataFrame,
    shipment_columns: list[str],
    cost_blocks: list[CostBlock],
    *,
    highlight_min_gt_max: bool = False,
    highlight_duplicate_lanes: bool = True,
    duplicate_lane_columns: tuple[str, ...] | None = None,
    highlight_missing_freight_block: str | None = None,
    transport_mode: str | None = None,
    bold_shipment_columns: frozenset[str] | None = None,
    supplier_name_column: str = "Supplier name",
) -> None:
    shipment_count = len(shipment_columns)
    bold_columns = bold_shipment_columns or frozenset()

    for col_idx, header in enumerate(shipment_columns, start=1):
        cell = ws.cell(row=COLUMN_HEADER_ROW, column=col_idx, value=header)
        style_header_cell(cell, bold=header in bold_columns)

    cost_col = shipment_count + 1
    for block in cost_blocks:
        write_cost_block_headers(ws, block, cost_col)
        cost_col += block_width(block)

    for row_idx in range(len(shipment_df)):
        excel_row = DATA_START_ROW + row_idx
        ocean_row = ocean_df.iloc[row_idx]
        shipment_row = shipment_df.iloc[row_idx]

        for col_idx, header in enumerate(shipment_columns, start=1):
            value = shipment_row[header]
            cell = ws.cell(
                row=excel_row,
                column=col_idx,
                value=format_shipment_cell_value(value),
            )
            apply_two_decimal_number_format(cell)

        cost_col = shipment_count + 1
        freight_has_values = False
        for block in cost_blocks:
            if should_fill_block(
                ocean_row,
                block,
                transport_mode=transport_mode,
                shipment_row=shipment_row,
                supplier_name_column=supplier_name_column,
            ):
                has_values = block_has_any_source_value(ocean_row, block)
                if highlight_missing_freight_block and block.title == highlight_missing_freight_block:
                    freight_has_values = has_values

                value_offset = 1 if block.include_currency else 0

                if block.include_currency and has_values:
                    currency = cell_text(ocean_row.get(CURRENCY_COLUMN))
                    if currency:
                        ws.cell(row=excel_row, column=cost_col, value=currency)

                if has_values and block.fill_when_min_equals_max and len(block.value_columns) == 1:
                    _, max_value, _ = pre_carriage_min_max_values(ocean_row)
                    write_rate_value_cell(
                        ws,
                        excel_row,
                        cost_col + value_offset,
                        format_cell_value(max_value),
                    )
                elif has_values and block.fill_when_min_not_equals_max and len(block.value_columns) == 3:
                    min_value, max_value, punit_value = pre_carriage_min_max_values(ocean_row)
                    write_rate_value_cell(
                        ws,
                        excel_row,
                        cost_col + value_offset,
                        format_cell_value(min_value),
                    )
                    write_rate_value_cell(
                        ws,
                        excel_row,
                        cost_col + value_offset + 1,
                        format_cell_value(punit_value),
                    )
                    write_rate_value_cell(
                        ws,
                        excel_row,
                        cost_col + value_offset + 2,
                        format_cell_value(max_value),
                    )

                    if (
                        highlight_min_gt_max
                        and min_value is not None
                        and max_value is not None
                        and min_value > max_value
                    ):
                        ws.cell(row=excel_row, column=cost_col + value_offset).fill = BLUE_FILL
                        ws.cell(row=excel_row, column=cost_col + value_offset + 2).fill = BLUE_FILL
                elif has_values:
                    for offset, value_column in enumerate(block.value_columns, start=value_offset):
                        source_column = value_column.source_column
                        if source_column is None:
                            continue
                        raw_value = ocean_row.get(source_column)
                        if value_column.percent_over_costs:
                            formatted = format_percent_over_costs_value(raw_value)
                        else:
                            formatted = format_cell_value(raw_value)
                        write_rate_value_cell(
                            ws,
                            excel_row,
                            cost_col + offset,
                            formatted,
                        )

            cost_col += block_width(block)

        if highlight_missing_freight_block and not freight_has_values:
            for col_idx in range(1, cost_col):
                ws.cell(row=excel_row, column=col_idx).fill = RED_FILL

    last_data_column = cost_col - 1
    if highlight_duplicate_lanes and last_data_column >= 1:
        lane_columns = duplicate_lane_columns
        if lane_columns is None and bold_columns:
            lane_columns = tuple(
                header for header in shipment_columns if header in bold_columns
            )
        if lane_columns:
            highlight_fully_duplicate_lane_rows(
                ws,
                shipment_columns=shipment_columns,
                last_data_column=last_data_column,
                duplicate_lane_columns=lane_columns,
                fill=GREEN_FILL,
            )

    for col_idx in range(1, cost_col):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18


def save_output_sheet(
    sheet_name: str,
    write_sheet,
    output_path: Path | None = None,
) -> Path:
    ensure_workspace_dirs()
    path = output_path or output_workbook_path()

    if path.exists():
        workbook = load_workbook(path)
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]
        worksheet = workbook.create_sheet(sheet_name)
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name

    write_sheet(worksheet)
    workbook.save(path)
    return path
