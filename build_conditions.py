"""Conditions tab and carrier/supplier name normalization rules."""

from __future__ import annotations

from dataclasses import dataclass

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from rate_layout_common import COLUMN_HEADER_ROW, DATA_START_ROW

CONDITIONS_SHEET = "Conditions"

CONDITIONS_COLUMNS = [
    "Column name",
    "RA name",
    "Name",
    "Operator",
    "Values",
]

SUPPLIER_NAME_COLUMN = "Supplier name"
CARRIER_NAME_COLUMN = "Carrier Name"

@dataclass(frozen=True)
class ConditionTarget:
    column_name: str
    tab_name: str


CONDITION_TARGETS: tuple[ConditionTarget, ...] = (
    ConditionTarget(SUPPLIER_NAME_COLUMN, "Sea Rates"),
    ConditionTarget(SUPPLIER_NAME_COLUMN, "Air MAWB rate"),
    ConditionTarget(SUPPLIER_NAME_COLUMN, "Air HAWB rate"),
    ConditionTarget(CARRIER_NAME_COLUMN, "Road pre-carriage for Sea"),
    ConditionTarget(CARRIER_NAME_COLUMN, "Road pre-carriage for Air MAWB"),
    ConditionTarget(CARRIER_NAME_COLUMN, "Road pre-carriage for Air HAWB"),
)

HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
CONDITION_APPLIED_FILL = PatternFill("solid", fgColor="D9D9D9")
BOLD = Font(bold=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


@dataclass(frozen=True)
class ConditionRule:
    name: str
    operator: str
    values: tuple[str, ...]

    def values_display(self) -> str:
        return ", ".join(f"'{value}'" for value in self.values)


CONDITION_RULES: tuple[ConditionRule, ...] = (
    ConditionRule(
        "ARX",
        "equals",
        ("ARX AE", "ARX SG", "ARX US", "ARX CN", "ARX HK", "ARX IE"),
    ),
    ConditionRule("RLWUS/RLWV", "equals", ("RLWUS", "RLWV")),
    ConditionRule("Bluewater CN/Bluewater US", "equals", ("Bluewater CN", "Bluewater US")),
    ConditionRule("DHL CN/DHL UK", "equals", ("DHL CN", "DHL UK")),
    ConditionRule("DHL UK/DHL UX", "equals", ("DHL UK", "DHL UX")),
    ConditionRule("EI IAH/EI NL", "equals", ("EI IAH", "EI NL")),
    ConditionRule("EI AE Int/EI IAH", "equals", ("EI IAH", "EI AE Int")),
    ConditionRule("Geodis AE/Geodis NL", "equals", ("Geodis AE", "Geodis NL")),
    ConditionRule("Geodis CN/Geodis NL", "equals", ("Geodis CN", "Geodis NL")),
    ConditionRule("Geodis NL/Geodis SG", "equals", ("Geodis NL", "Geodis SG")),
    ConditionRule("RLWV SG/RLWV", "equals", ("RLWV", "RLWV SG")),
    ConditionRule("DSV US/DSV CN", "equals", ("DSV US", "DSV CN")),
)


def lookup_replacement_name(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None

    for rule in CONDITION_RULES:
        if rule.operator.lower() == "equals" and text in rule.values:
            return rule.name
    return None


def normalize_condition_value(value: object) -> object:
    if value is None:
        return value

    text = str(value).strip()
    if not text:
        return value

    direct = lookup_replacement_name(text)
    if direct is not None:
        return direct

    if ";" not in text:
        return value

    parts = [part.strip() for part in text.split(";") if part.strip()]
    if not parts:
        return value

    replaced = [lookup_replacement_name(part) or part for part in parts]
    unique = list(dict.fromkeys(replaced))
    if len(unique) == 1:
        return unique[0]
    return "; ".join(unique)


def build_conditions_dataframe() -> pd.DataFrame:
    rows: list[dict[str, str]] = []
    for target in CONDITION_TARGETS:
        for rule in CONDITION_RULES:
            rows.append(
                {
                    "Column name": target.column_name,
                    "RA name": target.tab_name,
                    "Name": rule.name,
                    "Operator": rule.operator,
                    "Values": rule.values_display(),
                }
            )
    return pd.DataFrame(rows, columns=CONDITIONS_COLUMNS)


def _find_header_column(worksheet, header_name: str) -> int | None:
    for col_idx in range(1, worksheet.max_column + 1):
        if worksheet.cell(row=COLUMN_HEADER_ROW, column=col_idx).value == header_name:
            return col_idx
    return None


def _style_conditions_header(cell) -> None:
    cell.font = BOLD
    cell.fill = HEADER_FILL
    cell.alignment = LEFT


def _write_conditions_sheet(worksheet, conditions_df: pd.DataFrame) -> None:
    for col_idx, header in enumerate(CONDITIONS_COLUMNS, start=1):
        cell = worksheet.cell(row=1, column=col_idx, value=header)
        _style_conditions_header(cell)

    for row_offset, row in enumerate(conditions_df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            worksheet.cell(row=row_offset, column=col_idx, value=value)

    for col_idx in range(1, len(CONDITIONS_COLUMNS) + 1):
        worksheet.column_dimensions[get_column_letter(col_idx)].width = 28


def _style_condition_applied_cell(cell) -> None:
    cell.fill = CONDITION_APPLIED_FILL
    existing_font = cell.font
    cell.font = Font(
        name=existing_font.name,
        size=existing_font.size,
        bold=existing_font.bold,
        italic=existing_font.italic,
        color=existing_font.color,
        underline="single",
    )


def _apply_conditions_to_sheet(worksheet, target: ConditionTarget) -> int:
    column_idx = _find_header_column(worksheet, target.column_name)
    if column_idx is None:
        return 0

    replacements = 0
    for row_idx in range(DATA_START_ROW, worksheet.max_row + 1):
        cell = worksheet.cell(row=row_idx, column=column_idx)
        normalized = normalize_condition_value(cell.value)
        if normalized != cell.value:
            cell.value = normalized
            _style_condition_applied_cell(cell)
            replacements += 1
    return replacements


def apply_conditions_to_workbook(output_path) -> int:
    workbook = load_workbook(output_path)

    total_replacements = 0
    for target in CONDITION_TARGETS:
        if target.tab_name in workbook.sheetnames:
            total_replacements += _apply_conditions_to_sheet(workbook[target.tab_name], target)

    conditions_df = build_conditions_dataframe()
    if CONDITIONS_SHEET in workbook.sheetnames:
        del workbook[CONDITIONS_SHEET]
    conditions_ws = workbook.create_sheet(CONDITIONS_SHEET)
    _write_conditions_sheet(conditions_ws, conditions_df)

    workbook.save(output_path)
    return total_replacements


def finalize_output_workbook(output_path) -> None:
    replacements = apply_conditions_to_workbook(output_path)
    print(
        f"Built '{CONDITIONS_SHEET}' with {len(build_conditions_dataframe())} rows; "
        f"applied {replacements} condition replacement(s) -> {output_path}"
    )
