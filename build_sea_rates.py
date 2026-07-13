"""Build the Sea Rates output tab from the cleaned Ocean processing data."""

from __future__ import annotations

import re
import sys
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from carrier_groups import (
    CARRIER_GROUP_HOU,
    CARRIER_GROUP_MEA,
    CARRIER_GROUP_RTM,
    CARRIER_GROUP_SC,
    carrier_group_for_origin,
    is_arx_carrier_supplier_name,
)
from project_paths import INPUT_DIR, OUTPUT_DIR, PROCESSING_DIR, ensure_workspace_dirs
from rate_layout_common import apply_grouped_second_column_fill
from supplier_name_lookup import lookup_fred_supplier_name, map_fred_supplier_names

SEA_RATES_SHEET = "Sea Rates"
OCEAN_SOURCE_SHEET = "Ocean"
PROCESSING_GLOB = "combined_*.xlsx"
EXCEL_SUFFIXES = {".xlsx", ".xls", ".xlsm"}

COST_NAME_ROW = 1
APPLY_IF_ROW = 2
RATE_BY_ROW = 3
COLUMN_HEADER_ROW = 4
DATA_START_ROW = 5

HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
BOLD = Font(bold=True)
NORMAL = Font()
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

SERVICE_TYPE_MAP = {
    "LCL": "LTL/LCL",
    "FCL": "FCL/Multi_Load",
}

SEA_RATES_SHIPMENT_COLUMNS = [
    "RC name",
    "Scope of work",
    "Service Type",
    "Service type",
    "Lane ID",
    "Origin Country Code",
    "Origin Base City",
    "Port of loading",
    "Port of loading (second)",
    "Departure Seaport city",
    "Departure seaport name",
    "Destination country code",
    "Destination Seaport city",
    "Destination base city",
    "Destination base city (second)",
    "Destination Port code",
    "Supplier name",
    "Supplier code",
    "Valid From",
    "Valid to",
]

SEA_BOLD_SHIPMENT_COLUMNS = frozenset(
    {
        "Service type",
        "Origin Country Code",
        "Origin Base City",
        "Port of loading (second)",
        "Destination country code",
        "Destination base city (second)",
        "Destination Port code",
        "Supplier name",
        "Valid From",
        "Valid to",
    }
)

SEA_DEST_BASE_CITY_GROUP_COLUMNS = (
    "Service type",
    "Origin Country Code",
    "Origin Base City",
    "Port of loading (second)",
    "Destination country code",
    "Destination Port code",
    "Supplier name",
    "Valid From",
    "Valid to",
)

CURRENCY_COLUMN = "Quoting Currency"
EXPORT_CUSTOM_CLEARANCE_COL = "Export Custom Clearance (per BL)"
CERTIFICATE_OF_ORIGIN_COL = "Certificate of Origin (per BL)"


@dataclass(frozen=True)
class CostValueColumn:
    header: str
    source_column: str


@dataclass(frozen=True)
class CostBlock:
    title: str
    apply_if: str
    rate_by: str
    value_columns: tuple[CostValueColumn, ...]
    match_equipment_type: str | None = None
    match_service_type: str | None = None
    match_service_types: tuple[str, ...] | None = None
    match_carrier_groups: tuple[str, ...] | None = None
    match_arx_ae_lane: bool = False
    exclude_arx_ae_lane: bool = False


def _cell_text(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def _format_date(value: object) -> str:
    if pd.isna(value):
        return ""
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return _cell_text(value)
    return parsed.strftime("%d.%m.%Y")


def _trim_lane_id(lane_uid: object) -> str:
    text = _cell_text(lane_uid)
    if len(text) <= 4:
        return ""
    return text[:-4]


def _map_service_type(service_type: object) -> str:
    text = _cell_text(service_type)
    return SERVICE_TYPE_MAP.get(text.upper(), text)


def _rate_value(value: object) -> float | None:
    if pd.isna(value):
        return None
    text = _cell_text(value)
    if not text or text.lower() in {"on request", "n/a", "#n/a"}:
        return None
    try:
        return float(str(value).replace(",", "."))
    except (TypeError, ValueError):
        return None


def _format_cell_value(value: object) -> object:
    number = _rate_value(value)
    if number is None:
        return None
    if float(number).is_integer():
        return int(number)
    return number


def _output_equipment_type(equipment_type: str) -> str:
    text = _cell_text(equipment_type)
    if text.upper().endswith("DC"):
        return f"{text[:-2]}FT"
    return text


def _equipment_rate_by(equipment_type: str) -> str:
    if equipment_type.upper() == "LCL":
        return "Rate by: W/M"
    return "Rate by: Per container"


def _sorted_equipment_types(ocean_df: pd.DataFrame) -> list[str]:
    equipment_types = {
        _output_equipment_type(value)
        for value in ocean_df["Equipment Type"].dropna().tolist()
        if _cell_text(value)
    }
    return sorted(equipment_types, key=lambda value: (value.upper() == "LCL", value))


def list_input_files() -> list[Path]:
    return [
        path
        for path in sorted(INPUT_DIR.iterdir())
        if path.is_file()
        and path.suffix.lower() in EXCEL_SUFFIXES
        and not path.name.startswith("~$")
    ]


def get_rate_card_name() -> str:
    files = list_input_files()
    if not files:
        raise FileNotFoundError(f"No Excel files found in: {INPUT_DIR}")
    if len(files) == 1:
        return files[0].stem
    return files[0].stem


def latest_processing_file() -> Path:
    files = sorted(PROCESSING_DIR.glob(PROCESSING_GLOB))
    if not files:
        raise FileNotFoundError(
            f"No processing workbook found in {PROCESSING_DIR}. Run process_tabs.py first."
        )
    return files[-1]


def load_ocean_dataframe(processing_path: Path | None = None) -> pd.DataFrame:
    path = processing_path or latest_processing_file()
    workbook = pd.ExcelFile(path)
    sheet_name = next(
        (name for name in workbook.sheet_names if name.lower() == OCEAN_SOURCE_SHEET.lower()),
        None,
    )
    if sheet_name is None:
        raise ValueError(f"Sheet '{OCEAN_SOURCE_SHEET}' not found in {path.name}")
    return pd.read_excel(path, sheet_name=sheet_name)


def build_sea_rates_shipment_df(
    ocean_df: pd.DataFrame,
    rc_name: str | None = None,
) -> pd.DataFrame:
    rate_card_name = rc_name or get_rate_card_name()
    empty = pd.Series([""] * len(ocean_df), index=ocean_df.index, dtype="object")

    return pd.DataFrame(
        {
            "RC name": rate_card_name,
            "Scope of work": ocean_df["Scope of Work"],
            "Service Type": ocean_df["Service Type"],
            "Service type": ocean_df["Service Type"].map(_map_service_type),
            "Lane ID": ocean_df["Lane UID"].map(_trim_lane_id),
            "Origin Country Code": ocean_df["Origin Country Code"],
            "Origin Base City": ocean_df["Origin Base City"],
            "Port of loading": ocean_df["Port of loading POL (UNLOCODE)"],
            "Port of loading (second)": empty,
            "Departure Seaport city": ocean_df["Departure Seaport City"],
            "Departure seaport name": ocean_df["Departure Port Name"],
            "Destination country code": ocean_df["Destination Country Code"],
            "Destination Seaport city": ocean_df["Destination Seaport City"],
            "Destination base city": ocean_df["Destination Base City"],
            "Destination base city (second)": empty,
            "Destination Port code": ocean_df["Port of discharge POD (UNLOCODE)"],
            "Supplier name": map_fred_supplier_names(
                ocean_df["Origin Country Code"],
                ocean_df["Supplier"],
                transport_mode="ocean",
            ),
            "Supplier code": ocean_df["Supplier"],
            "Valid From": ocean_df["Valid from"].map(_format_date),
            "Valid to": ocean_df["Valid to"].map(_format_date),
        },
        columns=SEA_RATES_SHIPMENT_COLUMNS,
    )


def build_equipment_cost_blocks(equipment_types: list[str]) -> list[CostBlock]:
    blocks: list[CostBlock] = []

    for equipment_type in equipment_types:
        apply_if = f"Apply if: Equipment type equals '{equipment_type}'"
        rate_by = _equipment_rate_by(equipment_type)

        blocks.append(
            CostBlock(
                title=f"Freight Sea Intl ({equipment_type})",
                apply_if=apply_if,
                rate_by=rate_by,
                value_columns=(
                    CostValueColumn("MIN", "LCL min Ocean Freight Rate"),
                    CostValueColumn("p/unit", "Ocean Freight Rate (per ctnr or W/M)"),
                ),
                match_equipment_type=equipment_type,
            )
        )
        blocks.append(
            CostBlock(
                title=f"Sea Stuffing ({equipment_type})",
                apply_if=apply_if,
                rate_by=rate_by,
                value_columns=(
                    CostValueColumn(
                        "p/unit",
                        "Stuffing per container/consolidation fee per W/M",
                    ),
                ),
                match_equipment_type=equipment_type,
            )
        )
        blocks.append(
            CostBlock(
                title=f"Fumigation ({equipment_type})",
                apply_if=apply_if,
                rate_by=rate_by,
                value_columns=(
                    CostValueColumn("p/unit", "Container Fumigation [per Container]"),
                ),
                match_equipment_type=equipment_type,
            )
        )
        blocks.append(
            CostBlock(
                title=f"Sea DG surcharge ({equipment_type})",
                apply_if=apply_if,
                rate_by=rate_by,
                value_columns=(
                    CostValueColumn("p/unit", "IMO/DG Surcharge (per Container)"),
                ),
                match_equipment_type=equipment_type,
            )
        )

    return blocks


def build_service_cost_blocks() -> list[CostBlock]:
    blocks: list[CostBlock] = []

    for service_type in ("LCL", "FCL"):
        apply_if = f"Apply if: Service type equals '{service_type}'"
        blocks.append(
            CostBlock(
                title=f"Sea Waiver Fee ({service_type})",
                apply_if=apply_if,
                rate_by="Rate by: Per BL",
                value_columns=(CostValueColumn("p/unit", "Waiver Fee (per BL)"),),
                match_service_type=service_type,
            )
        )
        blocks.append(
            CostBlock(
                title=f"Sea OBL Fee ({service_type})",
                apply_if=apply_if,
                rate_by="Rate by: Per BL",
                value_columns=(CostValueColumn("p/unit", "OBL Fee (per BL)"),),
                match_service_type=service_type,
            )
        )

    return blocks


def build_carrier_group_cost_blocks() -> list[CostBlock]:
    blocks: list[CostBlock] = []
    carrier_group_specs = (
        (
            "Sea Certificate of origin (LCL, FCL HOU)",
            CERTIFICATE_OF_ORIGIN_COL,
            (CARRIER_GROUP_HOU,),
            False,
            False,
        ),
        (
            "Sea Export clearance (LCL, FCL HOU)",
            EXPORT_CUSTOM_CLEARANCE_COL,
            (CARRIER_GROUP_HOU,),
            False,
            True,
        ),
        (
            "Sea Certificate of origin (LCL, FCL MEA, SC, RTM)",
            CERTIFICATE_OF_ORIGIN_COL,
            (CARRIER_GROUP_MEA, CARRIER_GROUP_SC, CARRIER_GROUP_RTM),
            False,
            False,
        ),
        (
            "Sea Export clearance (LCL, FCL MEA, SC, RTM)",
            EXPORT_CUSTOM_CLEARANCE_COL,
            (CARRIER_GROUP_MEA, CARRIER_GROUP_SC, CARRIER_GROUP_RTM),
            False,
            True,
        ),
        (
            "Sea Export clearance (ARX AE)",
            EXPORT_CUSTOM_CLEARANCE_COL,
            (),
            True,
            False,
        ),
    )

    for title, source_column, carrier_groups, arx_ae_lane, exclude_arx_ae_lane in carrier_group_specs:
        if arx_ae_lane:
            apply_if = (
                "Apply if: Service type in {'LCL', 'FCL'} "
                "and Supplier name equals 'ARX'"
            )
        elif carrier_groups == (CARRIER_GROUP_HOU,):
            apply_if = (
                "Apply if: Service type in {'LCL', 'FCL'} "
                "and carrier group equals 'HOU'"
            )
        else:
            apply_if = (
                "Apply if: Service type in {'LCL', 'FCL'} "
                "and carrier group in {'MEA', 'SC', 'RTM'}"
            )

        blocks.append(
            CostBlock(
                title=title,
                apply_if=apply_if,
                rate_by="Rate by: Per BL",
                value_columns=(CostValueColumn("p/unit", source_column),),
                match_service_types=("LCL", "FCL"),
                match_carrier_groups=carrier_groups or None,
                match_arx_ae_lane=arx_ae_lane,
                exclude_arx_ae_lane=exclude_arx_ae_lane,
            )
        )

    return blocks


def build_sea_rates_cost_blocks(ocean_df: pd.DataFrame) -> list[CostBlock]:
    equipment_types = _sorted_equipment_types(ocean_df)
    return (
        build_equipment_cost_blocks(equipment_types)
        + build_service_cost_blocks()
        + build_carrier_group_cost_blocks()
    )


def _block_width(block: CostBlock) -> int:
    return 1 + len(block.value_columns)


def _is_arx_carrier_lane(ocean_row: pd.Series) -> bool:
    supplier_name = lookup_fred_supplier_name(
        ocean_row.get("Origin Country Code"),
        ocean_row.get("Supplier"),
        transport_mode="ocean",
    )
    return is_arx_carrier_supplier_name(supplier_name)


def _row_matches_block(ocean_row: pd.Series, block: CostBlock) -> bool:
    if block.match_equipment_type is not None:
        source_equipment = _cell_text(ocean_row["Equipment Type"])
        if _output_equipment_type(source_equipment) != block.match_equipment_type:
            return False
    if block.match_service_types is not None:
        service_type = _cell_text(ocean_row["Service Type"]).upper()
        allowed = {value.upper() for value in block.match_service_types}
        if service_type not in allowed:
            return False
    elif block.match_service_type is not None:
        if _cell_text(ocean_row["Service Type"]).upper() != block.match_service_type.upper():
            return False

    arx_carrier_lane = _is_arx_carrier_lane(ocean_row)
    if block.match_arx_ae_lane:
        return arx_carrier_lane
    if block.exclude_arx_ae_lane and arx_carrier_lane:
        return False
    if block.match_carrier_groups is not None:
        carrier_group = carrier_group_for_origin(ocean_row.get("Origin Country Code"))
        if carrier_group not in block.match_carrier_groups:
            return False

    if (
        block.match_equipment_type is None
        and block.match_service_type is None
        and block.match_service_types is None
        and block.match_carrier_groups is None
        and not block.match_arx_ae_lane
    ):
        return False
    return True


def _style_header_cell(cell, *, bold: bool = False, center: bool = False) -> None:
    cell.font = BOLD if bold else NORMAL
    cell.fill = HEADER_FILL
    cell.alignment = CENTER if center else LEFT


def _write_merged_header_row(
    ws: Worksheet,
    row_index: int,
    start_col: int,
    width: int,
    value: str,
) -> None:
    end_col = start_col + width - 1
    if width > 1:
        ws.merge_cells(
            start_row=row_index,
            start_column=start_col,
            end_row=row_index,
            end_column=end_col,
        )
    cell = ws.cell(row=row_index, column=start_col, value=value)
    _style_header_cell(cell)


def _write_cost_block_headers(ws: Worksheet, block: CostBlock, start_col: int) -> None:
    width = _block_width(block)

    _write_merged_header_row(ws, COST_NAME_ROW, start_col, width, block.title)
    _write_merged_header_row(ws, APPLY_IF_ROW, start_col, width, block.apply_if)
    _write_merged_header_row(ws, RATE_BY_ROW, start_col, width, block.rate_by)

    currency_cell = ws.cell(row=COLUMN_HEADER_ROW, column=start_col, value="Currency")
    _style_header_cell(currency_cell, center=True)

    for offset, value_column in enumerate(block.value_columns, start=1):
        cell = ws.cell(row=COLUMN_HEADER_ROW, column=start_col + offset, value=value_column.header)
        _style_header_cell(cell, center=True)


def write_sea_rates_sheet(
    ws: Worksheet,
    ocean_df: pd.DataFrame,
    shipment_df: pd.DataFrame,
    cost_blocks: list[CostBlock],
) -> None:
    shipment_count = len(SEA_RATES_SHIPMENT_COLUMNS)

    for col_idx, header in enumerate(SEA_RATES_SHIPMENT_COLUMNS, start=1):
        cell = ws.cell(row=COLUMN_HEADER_ROW, column=col_idx, value=header)
        _style_header_cell(cell, bold=header in SEA_BOLD_SHIPMENT_COLUMNS)

    cost_col = shipment_count + 1
    for block in cost_blocks:
        _write_cost_block_headers(ws, block, cost_col)
        cost_col += _block_width(block)

    for row_idx in range(len(shipment_df)):
        excel_row = DATA_START_ROW + row_idx
        ocean_row = ocean_df.iloc[row_idx]
        shipment_row = shipment_df.iloc[row_idx]

        for col_idx, header in enumerate(SEA_RATES_SHIPMENT_COLUMNS, start=1):
            value = shipment_row[header]
            ws.cell(
                row=excel_row,
                column=col_idx,
                value=None if value == "" or pd.isna(value) else value,
            )

        cost_col = shipment_count + 1
        for block in cost_blocks:
            if _row_matches_block(ocean_row, block):
                currency = _cell_text(ocean_row.get(CURRENCY_COLUMN))
                if currency:
                    ws.cell(row=excel_row, column=cost_col, value=currency)

                for offset, value_column in enumerate(block.value_columns, start=1):
                    amount = _format_cell_value(ocean_row.get(value_column.source_column))
                    if amount is not None:
                        value_cell = ws.cell(row=excel_row, column=cost_col + offset, value=amount)
                        value_cell.number_format = "0.00##"

            cost_col += _block_width(block)

    for col_idx in range(1, cost_col):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18


def output_workbook_path() -> Path:
    rc_name = re.sub(r'[\\/*?:\[\]]+', "_", get_rate_card_name()).strip()
    return OUTPUT_DIR / f"{rc_name}_output.xlsx"


def save_sea_rates_workbook(
    ocean_df: pd.DataFrame,
    shipment_df: pd.DataFrame,
    cost_blocks: list[CostBlock],
    output_path: Path | None = None,
) -> Path:
    ensure_workspace_dirs()
    path = output_path or output_workbook_path()

    if path.exists():
        workbook = load_workbook(path)
        if SEA_RATES_SHEET in workbook.sheetnames:
            del workbook[SEA_RATES_SHEET]
        worksheet = workbook.create_sheet(SEA_RATES_SHEET)
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = SEA_RATES_SHEET

    write_sea_rates_sheet(worksheet, ocean_df, shipment_df, cost_blocks)
    workbook.save(path)
    return path


def build_sea_rates(
    processing_path: Path | None = None,
    rc_name: str | None = None,
    output_path: Path | None = None,
) -> Path:
    ocean_df = load_ocean_dataframe(processing_path)
    shipment_df = build_sea_rates_shipment_df(ocean_df, rc_name=rc_name)
    shipment_df = apply_grouped_second_column_fill(
        shipment_df,
        first_column="Destination base city",
        second_column="Destination base city (second)",
        group_columns=SEA_DEST_BASE_CITY_GROUP_COLUMNS,
    )
    cost_blocks = build_sea_rates_cost_blocks(ocean_df)
    saved_path = save_sea_rates_workbook(
        ocean_df,
        shipment_df,
        cost_blocks,
        output_path=output_path,
    )
    print(
        f"Built '{SEA_RATES_SHEET}' with {len(shipment_df)} rows and "
        f"{len(cost_blocks)} cost blocks -> {saved_path}"
    )
    return saved_path


def main() -> int:
    try:
        build_sea_rates()
        return 0
    except KeyboardInterrupt:
        print("\nOperation cancelled.")
        return 130
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
